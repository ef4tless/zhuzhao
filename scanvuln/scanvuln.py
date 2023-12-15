import os

import openpyxl
import pandas as pd
import re
import datetime
import win32com.client as win32
from DrissionPage import WebPage, ChromiumOptions, SessionOptions
from DrissionPage import ChromiumPage
from DrissionPage.common import Actions


# 设置参数
startnum = 1
company_province = '天津'
company_city = '天津'
discovery_date = "20231215"
cves = [
        {'cve': "CVE-2021-29441", 'describe': "Nacos默认配置未授权访问漏洞", "type": "配置错误"},
        {'cve': "CVE-2021-29442", 'describe': "nacos 关键功能的认证机制缺失", "type": "配置错误"},
        {'cve': "CVE-2019-11248", 'describe': "Kubelet信息泄漏漏洞", "type": "信息泄露"},
        {'cve': "CVE-2021-33544", 'describe': "Geutebrück G-Cam E2 命令注入漏洞", "type": "命令执行"},
        {'cve': "CVE-2022-25481", 'describe': "ThinkPHP 信息泄露", "type": "信息泄露"},
        {'cve': "CVE-2009-4223", 'describe': "KR-Web文件包含", "type": "其他"},
        {'cve': "CVE-2015-8813", 'describe': "Umbraco存在SSRF攻击", "type": "其他"},
        {'cve': "CVE-2018-1000600", 'describe': "Jenkins存在信息泄露", "type": "信息泄露"},
        {'cve': "CVE-2018-6910", 'describe': "DedeCMS信息泄露", "type": "信息泄露"},
        {'cve': "CVE-2021-4191", 'describe': "GitLab Enterprise Edition和GitLab Community Edition 授权问题漏洞", "type": "配置错误"},
        {'cve': "CVE-2017-18638", 'describe': "Graphite存在SSRF攻击", "type": "其他"},
        {'cve': "CVE-2018-15517", 'describe': "D-Link Central WiFiManager CWM-100存在SSRF攻击", "type": "其他"},
        {'cve': "CVE-2019-19824", 'describe': "TOTOLINK Realtek SDK 的路由器存在命令执行漏洞", "type": "命令执行"},
        {'cve': "CVE-2022-46463", 'describe': "关键功能的认证机制缺失", "type": "配置错误"},
        {'cve': "CVE-2020-11110", 'describe': "Grafana 跨站脚本漏洞", "type": "XSS"},
        {'cve': "CVE-2021-27358", 'describe': "Grafana拒绝服务漏洞", "type": "拒绝服务"},
        {'cve': "CVE-2020-26413", 'describe': "Gitlab graphql 敏感信息泄漏漏洞", "type": "信息泄露"},
        {'cve': "CVE-2010-0157", 'describe': "Joomla目录遍历漏洞", "type": "目录遍历"},
        {'cve': "CVE-2009-3318", 'describe': "Joomla目录遍历漏洞", "type": "目录遍历"},
        {'cve': "CVE-2009-4679", 'describe': "Joomla文件包含", "type": "其他"},
        {'cve': "CVE-2009-4202", 'describe': "Joomla目录遍历漏洞", "type": "目录遍历"},
        {'cve': "CVE-2011-3315", 'describe': "Cisco Unified Contact Center Express 目录遍历漏洞", "type": "目录遍历"},
        {'cve': "CVE-2010-2918", 'describe': "Joomla命令执行漏洞", "type": "命令执行"},
        {'cve': "CVE-2012-1226", 'describe': "Alpha 中存在多个目录遍历漏洞", "type": "目录遍历"},
        {'cve': "CVE-2015-1000005", 'describe': "wordpress插件远程文件下载漏洞", "type": "信息泄露"},
        {'cve': "CVE-2015-2996", 'describe': "SysAid Help Desk 中存在多个目录遍历漏洞", "type": "目录遍历"},
        {'cve': "CVE-2016-2389", 'describe': "SAP NetWeaver存在目录遍历漏洞", "type": "目录遍历"},
        {'cve': "CVE-2021-3577", 'describe': "Binatone Hubble 存在命令执行漏洞", "type": "命令执行"},
        {'cve': "CVE-2018-12613", 'describe': "phpMyAdmin存在代码执行漏洞", "type": "命令执行"},
        {'cve': "CVE-2018-16299", 'describe': "WordPress存在目录遍历", "type": "目录遍历"},
        {'cve': "CVE-2018-18323", 'describe': "CentOS Web Panel 存在目录遍历漏洞", "type": "目录遍历"},
        {'cve': "CVE-2018-19753", 'describe': "Tarantella Enterprise存在目录遍历漏洞", "type": "目录遍历"},
        {'cve': "CVE-2018-6008", 'describe': "Joomla任意文件下载漏洞", "type": "信息泄露"},
        {'cve': "CVE-2019-16123", 'describe': "Kartatopia PilusCart存在文件泄露漏洞", "type": "信息泄露"},
        {'cve': "CVE-2021-39433", 'describe': "BIQS IT Biqs-drive存在任意文件读取漏洞", "type": "信息泄露"},
        {'cve': "CVE-2021-39316", 'describe': "WordPress 的 Zoomsounds 插件存在目录遍历漏洞", "type": "目录遍历"},
        {'cve': "CVE-2021-41291", 'describe': "ECOA BAS 控制器存在信息泄露漏洞", "type": "信息泄露"},
        {'cve': "CVE-2022-24900", 'describe': "Piano LED Visualizer存在目录遍历漏洞", "type": "目录遍历"},
        {'cve': "CVE-2019-2616", 'describe': "Oracle Fusion Middleware存在命令执行漏洞", "type": "命令执行"},
        {'cve': "CVE-2019-2767", 'describe': "Oracle Fusion Middleware存在命令执行漏洞", "type": "命令执行"},
        {'cve': "CVE-2019-20224", 'describe': "Pandora FMS存在命令执行漏洞", "type": "命令执行"},
        {'cve': "CVE-2021-24917", 'describe': "WPS存在登录绕过漏洞", "type": "其他"},
        {'cve': "CVE-2021-37304", 'describe': "jeecg-boot存在信息泄露漏洞", "type": "信息泄露"},
        {'cve': "CVE-2021-37305", 'describe': "jeecg-boot存在信息泄露漏洞", "type": "信息泄露"},
        {'cve': "CVE-2022-2185", 'describe': "GitLab存在命令执行漏洞", "type": "命令执行"},
        {'cve': "CVE-2021-43734", 'describe': "kkFileview存在目录遍历漏洞", "type": "目录遍历"},
        {'cve': "CNVD-2021-15824", 'describe': "EmpireCMS存在XSS漏洞", "type": "XSS"},
        {'cve': "CVE-2022-26233", 'describe': "通过Suite 2.9 Build 0275的Barco Control Room Management存在目录遍历", "type": "目录遍历"},
        {'cve': "CVE-2021-21311", 'describe': "adminer adminer 服务端请求伪造", "type": "其他"},
        {'cve': "CVE-2019-8442", 'describe': "Atlassian Jira Webroot 目录遍历漏洞", "type": "目录遍历"},
        {'cve': "CVE-2018-7719", 'describe': "Acrolinx Server for Windows路径遍历漏洞", "type": "目录遍历"},
        {'cve': "CVE-2017-10271", 'describe': "WebLogic XMLDecoder反序列化漏洞", "type": "其他"},
        {'cve': "CVE-2021-21311", 'describe': "adminer adminer 服务端请求伪造", "type": "其他"},
        {'cve': "CVE-2021-40822", 'describe': "GeoServer TestWfsPost SSRF漏洞", "type": "其他"},
        {'cve': "CVE-2021-41460", 'describe': "shopex ecshop sql命令中使用的特殊元素转义处理不恰当", "type": "SQL注入"},
        {'cve': "CNVD-2022-86535", 'describe': "ThinkPHP命令执行漏洞", "type": "命令执行"},
        {'cve': "CVE-2022-36883", 'describe': "jenkins git 授权机制缺失", "type": "配置错误"},
        {'cve': "CVE-2022-21371", 'describe': "WebLogic Server 目录遍历漏洞", "type": "目录遍历"},
        {'cve': "CVE-2018-8033", 'describe': "Apache OFBiz xmlrpc XXE漏洞", "type": "其他"},
        {'cve': "CVE-2023-29887", 'describe': "spreadsheet-reader存在文件包含漏洞", "type": "其他"},
        {'cve': "CNVD-2021-28277", 'describe': "蓝凌 oa 前台任意文件读取漏洞", "type": "信息泄露"},
        {'cve': "CVE-2017-3506", 'describe': "Oracle WebLogic Server XMLDecoder 反序列化漏洞", "type": "命令执行"},
        # {'cve': "", 'describe': "", "type": ""}
]

print("[+]开始执行高危漏洞数据处理......")
with open('result.txt', 'r') as file:
    scan_results = file.read()

pattern = r"\[(CNVD-\d{4}-\d{5}|CVE-\d{4}-\d{5})?\] \[\w+\] \[high\] ((?:http|https)://[^:]+):(\d+)"
parsed_results = re.findall(pattern, scan_results)

def process_province_name(name):
    if not isinstance(name, str):
        return ""
    for suffix in ["市", "省", "维吾尔自治区", "自治区", "州", "特别市"]:
        if name.endswith(suffix):
            return name.replace(suffix, "")
    return name



cve_dict = {item['cve']: {'describe': item['describe'], 'type': item['type']} for item in cves}

province_code_map = {
    "北京": 1, "天津": 2, "河北": 3, "山西": 4, "内蒙古": 5, "辽宁": 6, "吉林": 7, "黑龙江": 8,
    "上海": 9, "江苏": 10, "浙江": 11, "安徽": 12, "福建": 13, "江西": 14, "山东": 15, "河南": 16,
    "湖北": 17, "湖南": 18, "广东": 19, "广西": 20, "海南": 21, "重庆": 22, "四川": 23, "贵州": 24,
    "云南": 25, "西藏": 26, "陕西": 27, "甘肃": 28, "青海": 29, "宁夏": 30, "新疆": 31, "兵团": 32
}

def generate_clue_number(province, date_str, sequence):
    province_code = province_code_map.get(province, "0")
    return f"{province_code:02d}02{date_str}{sequence:08d}"

info_df = pd.read_excel('info.xlsx')

final_data = []
for sequence, (cve, url, port) in enumerate(parsed_results, start=startnum):
    if cve not in cve_dict:
        print(cve)
        continue
    domain = url.split('//')[1].split('/')[0]
    protocol = 'http' if port == '80' else 'https' if port == '443' else '未知'
    matched_row = info_df[info_df['网站Host（域名）'] == domain].iloc[0]
    province = process_province_name(matched_row['省份（中文）'])
    vuln_name = cve_dict.get(cve, {}).get('describe', '')
    vuln_type = cve_dict.get(cve, {}).get('type', '')

    row = {
        '线索编号': generate_clue_number(company_province, discovery_date[2:], sequence),
        '系统域名': domain,
        'IP地址': matched_row['IP地址'],
        '服务端口': port,
        'URL': url,
        '协议': protocol,
        '漏洞名称': vuln_name,
        '漏洞编号': cve,
        '漏洞类型': vuln_type,
        'IP归属省份': province,
        '发现时间': discovery_date,
        '是否异地': '是',
        '所属单位': matched_row['ICP备案单位'],
        '关联个人姓名': '',
        '关联个人身份证号': '',
        '关联个人手机号': '',
        '所属省份': company_province,
        '所属地市': company_city,
        '所属区县': '',
        '服务类型': 'Web服务'
    }
    final_data.append(row)

final_df = pd.DataFrame(final_data)
final_df.to_excel('Vuln_output.xlsx', index=False)
print("[+]Vuln_output.xlsx已导出")

output_df = pd.read_excel('Vuln_output.xlsx')
companies = output_df['所属单位']

try:
    tyc_workbook = openpyxl.load_workbook('天眼查.xlsx')
except FileNotFoundError:
    tyc_workbook = openpyxl.Workbook()

if 'Sheet' in tyc_workbook.sheetnames:
    sheet = tyc_workbook['Sheet']
else:
    sheet = tyc_workbook.active

if sheet.max_row > 3:
    sheet.delete_rows(4, sheet.max_row - 2)

row_index = 4
for company in companies:
    if row_index > sheet.max_row:
        sheet.append([company])
    else:
        sheet.cell(row=row_index, column=1, value=company)
    row_index += 1


tyc_workbook.save('天眼查.xlsx')
excel_file_path = '天眼查.xlsx'
excel = win32.gencache.EnsureDispatch('Excel.Application')

excel.Visible = False
try:
    workbook = excel.Workbooks.Open(os.path.abspath(excel_file_path))
    workbook.Save()
    workbook.Close()
finally:
    excel.Quit()

del excel
print("[+]天眼查.xlsx已导出")
co = ChromiumOptions()
so = SessionOptions()
page = WebPage(driver_or_options=co, session_or_options=so)

page.get('https://www.tianyancha.com/batch')
ac = Actions(page)
ac.click('.btn btn-primary btn-upload -w88 -h36 -batch-upload')

