import re
import os
import openpyxl
import pandas as pd
import win32com.client as win32
from datetime import datetime
from DrissionPage import WebPage, ChromiumOptions, SessionOptions
from DrissionPage import ChromiumPage
from DrissionPage.common import Actions

# 需要设置的参数
startnum = 2098
company_city = "北京"
company_province = "北京"
discovery_date = '20231215'

print("[+]开始执行高危端口数据处理......")

def generate_clue_number(province_number, date_str, clue_index):
    return f"{province_number:02d}03{date_str}{clue_index:08d}"

def is_person_name(name):
    return bool(re.match(r'^[\u4e00-\u9fa5]{2,4}$', name))

def process_province_name(name):
    if not isinstance(name, str):
        return ""
    for suffix in ["市", "省", "维吾尔自治区", "自治区", "州", "特别市"]:
        if name.endswith(suffix):
            return name.replace(suffix, "")
    return name

province_mapping = {
    "北京": 1, "天津": 2, "河北": 3, "山西": 4, "内蒙古": 5, "辽宁": 6, "吉林": 7, "黑龙江": 8,
    "上海": 9, "江苏": 10, "浙江": 11, "安徽": 12, "福建": 13, "江西": 14, "山东": 15, "河南": 16,
    "湖北": 17, "湖南": 18, "广东": 19, "广西": 20, "海南": 21, "重庆": 22, "四川": 23, "贵州": 24,
    "云南": 25, "西藏": 26, "陕西": 27, "甘肃": 28, "青海": 29, "宁夏": 30, "新疆": 31, "兵团": 32
}

scan_results_df = pd.read_excel('scan_results.xlsx')
info_df = pd.read_excel('info.xlsx')
fields = [
    "ftp", "ssh", "telnet", "smtp", "dns", "tftp", "http", "pop3", "rpcbind", "nfs", "msrpc",
    "netbios-ns", "netbios-ssn", "microsoft-ds", "ldap", "rsync", "socks", "ms-sql-s", "squid-http",
    "mysql", "ms-wbt-server", "rdp", "remote_desktop", "remote_desktop_connection", "ibm-db2", "vnc",
    "redis", "apache-struts2", "mongod", "upnp", "oracle", "activemq"
]
new_df = pd.DataFrame(columns=[
    '线索编号', '系统域名', 'IP地址', '高危端口', 'IP发现省份', '发现时间', '是否异地',
    '所属单位', '关联个人姓名', '关联个人身份证号', '关联个人手机号', '所属省份', '所属地市',
    '所属区县', '服务类型'
])
clue_index = startnum

for index, row in scan_results_df.iterrows():
    fingerprint = row['指纹']
    if isinstance(fingerprint, str) and any(field in fingerprint and not (field == 'http' and 'https' in fingerprint) for field in fields):
        ip, port = row['主机地址'].split(':')
        info_match = info_df[info_df['IP地址'] == ip]
        if not info_match.empty:
            info_row = info_match.iloc[0]
            province = process_province_name(info_row['省份（中文）'])
            icp_unit = info_row['ICP备案单位']
            if not is_person_name(icp_unit):
                new_row = pd.DataFrame([{
                    '线索编号': generate_clue_number(province_mapping.get(company_province, 0), discovery_date[2:], clue_index), # province
                    '系统域名': info_row['网站Host（域名）'],
                    'IP地址': ip,
                    '高危端口': port,
                    'IP发现省份': province,
                    '发现时间': discovery_date,
                    '是否异地': '是',
                    '所属单位': info_row['ICP备案单位'],
                    '关联个人姓名': '',
                    '关联个人身份证号': '',
                    '关联个人手机号': '',
                    '所属省份': company_province,
                    '所属地市': company_city,
                    '所属区县': '',
                    '服务类型': '操作系统及通用软件服务'
                }])
                new_df = pd.concat([new_df, new_row], ignore_index=True)
                clue_index += 1

new_df.to_excel('Port_output.xlsx', index=False)
print("[+]Port_output.xlsx已导出")

output_df = pd.read_excel('Port_output.xlsx')

companies = output_df['所属单位']


try:
    tyc_workbook = openpyxl.load_workbook('天眼查.xlsx')
except FileNotFoundError:
    tyc_workbook = openpyxl.Workbook()

if 'Sheet' in tyc_workbook.sheetnames:
    sheet = tyc_workbook['Sheet']
else:
    sheet = tyc_workbook.active

if sheet.max_row > 3:
    sheet.delete_rows(4, sheet.max_row - 2)

row_index = 4
for company in companies:
    if row_index > sheet.max_row:
        sheet.append([company])
    else:
        sheet.cell(row=row_index, column=1, value=company)
    row_index += 1


tyc_workbook.save('天眼查.xlsx')

excel_file_path = '天眼查.xlsx'
excel = win32.gencache.EnsureDispatch('Excel.Application')

excel.Visible = False
try:
    workbook = excel.Workbooks.Open(os.path.abspath(excel_file_path))
    workbook.Save()
    workbook.Close()

finally:
    excel.Quit()

del excel
print("[+]天眼查.xlsx已导出")
co = ChromiumOptions()
so = SessionOptions()
page = WebPage(driver_or_options=co, session_or_options=so)

page.get('https://www.tianyancha.com/batch')
ac = Actions(page)
ac.click('.btn btn-primary btn-upload -w88 -h36 -batch-upload')



