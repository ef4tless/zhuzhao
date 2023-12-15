import os
import pandas as pd
import re
import openpyxl
import win32com.client as win32
from datetime import datetime
from DrissionPage import WebPage, ChromiumOptions, SessionOptions
from DrissionPage import ChromiumPage
from DrissionPage.common import Actions

# 请设置变量
startnum = 1
company_city = "北京"
company_province = "北京"
date = "20231214"

print("[+]开始执行弱口令数据处理......")
file_path = f"logs/{date[:4]}-{date[4:6]}-{date[6:]}/success.txt"
if not os.path.exists(file_path):
    print("文件不存在")
    exit()

with open(file_path, 'r') as file:
    lines = file.readlines()
data = []
for line in lines:
    match = re.search(r'http(s)?://([^:/]+):(\d+)', line)
    if match:
        protocol = 'https' if match.group(1) else 'http'
        domain = match.group(2)
        port = match.group(3)
        data.append([domain, port, protocol])

success_df = pd.DataFrame(data, columns=['系统域名', '服务端口', '协议'])
info_df = pd.read_excel('info.xlsx')

def process_province_name(name):
    if not isinstance(name, str):
        return ""
    for suffix in ["市", "省", "维吾尔自治区", "自治区", "州", "特别市"]:
        if name.endswith(suffix):
            return name.replace(suffix, "")
    return name

final_data = []
for _, row in success_df.iterrows():
    domain = row['系统域名']
    info_row = info_df[info_df['网站Host（域名）'] == domain]
    if not info_row.empty:
        ip_address = info_row['IP地址'].values[0]
        operator = info_row['运营商'].values[0]
        province = info_row['省份（中文）'].values[0]
        icp_unit = info_row['ICP备案单位'].values[0]

        final_data.append([domain, ip_address, operator, row['服务端口'], row['协议'], process_province_name(province), icp_unit])
    else:
        final_data.append([domain, None, None, row['服务端口'], row['协议'], None, None])

columns = ['系统域名', 'IP地址', '运营商', '服务端口', '协议', '省份（中文）', '所属单位']
final_df = pd.DataFrame(final_data, columns=columns)


def generate_clue_number(province, date, index):
    province_code = {
        "北京": "01", "天津": "02", "河北": "03", "山西": "04", "内蒙古": "05",
        "辽宁": "06", "吉林": "07", "黑龙江": "08", "上海": "09", "江苏": "10",
        "浙江": "11", "安徽": "12", "福建": "13", "江西": "14", "山东": "15",
        "河南": "16", "湖北": "17", "湖南": "18", "广东": "19", "广西": "20",
        "海南": "21", "重庆": "22", "四川": "23", "贵州": "24", "云南": "25",
        "西藏": "26", "陕西": "27", "甘肃": "28", "青海": "29", "宁夏": "30",
        "新疆": "31", "兵团": "32"
    }
    return province_code.get(province, "00") + "01" + date + f"{index:08d}"



custom_sequence = startnum
if '线索编号' not in final_df.columns:
    final_df['线索编号'] = pd.NA
for index, row in final_df.iterrows():
    clue_number = generate_clue_number(row['省份（中文）'], date[2:], custom_sequence)
    final_df.at[index, '线索编号'] = clue_number
    custom_sequence += 1  

final_df['https://sys.thefirst.ltd/login'] = 'http://' + final_df['系统域名'] + ':' + final_df['服务端口']
final_df['弱口令分类'] = '口令复杂度不够'
final_df['发现时间'] = date
final_df['是否异地'] = '是'
final_df['关联个人姓名'] = ''
final_df['关联个人身份证号'] = ''
final_df['关联个人手机号'] = ''
final_df['所属省份'] = company_province
final_df['所属地市'] = company_city
final_df['所属区县'] = ''
final_df['服务类型'] = 'Web服务'

columns_to_keep = ['线索编号', '系统域名', 'IP地址', '运营商', '服务端口', 'https://sys.thefirst.ltd/login', '协议',
                   '弱口令分类', '省份（中文）', '发现时间', '是否异地', '所属单位', '关联个人姓名',
                   '关联个人身份证号', '关联个人手机号', '所属省份', '所属地市', '所属区县', '服务类型']
final_df = final_df[columns_to_keep]

final_df.to_excel('Passwd_output.xlsx', index=False)

print("[+]Passwd_output.xlsx已导出")

output_df = pd.read_excel('Passwd_output.xlsx')

companies = output_df['所属单位']


try:
    tyc_workbook = openpyxl.load_workbook('天眼查.xlsx')
except FileNotFoundError:
    tyc_workbook = openpyxl.Workbook()

if 'Sheet' in tyc_workbook.sheetnames:
    sheet = tyc_workbook['Sheet']
else:
    sheet = tyc_workbook.active


if sheet.max_row > 3:
    sheet.delete_rows(4, sheet.max_row - 2)


row_index = 4 
for company in companies:
    if row_index > sheet.max_row:
        sheet.append([company])
    else:
        sheet.cell(row=row_index, column=1, value=company)
    row_index += 1


tyc_workbook.save('天眼查.xlsx')

excel_file_path = '天眼查.xlsx'
excel = win32.gencache.EnsureDispatch('Excel.Application')

excel.Visible = False
try:
    workbook = excel.Workbooks.Open(os.path.abspath(excel_file_path))
    workbook.Save()
    workbook.Close()

finally:
    excel.Quit()

del excel
print("[+]天眼查.xlsx已导出")
co = ChromiumOptions()
so = SessionOptions()
page = WebPage(driver_or_options=co, session_or_options=so)

page.get('https://www.tianyancha.com/batch')
ac = Actions(page)
ac.click('.btn btn-primary btn-upload -w88 -h36 -batch-upload')
